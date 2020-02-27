from openpyxl import Workbook, load_workbook
import xlwt

class excel_xlsx:
    arquivo_excel = ''
    arquivo_atual = ''
    def __init__(self):
        self.arquivo_excel = Workbook()
        self.arquivo_atual = self.arquivo_excel.active

    def escrever_na_celula(self, celula, valor):
        self.arquivo_atual[celula] = valor

    def add_dados(self, *dados):
        self.arquivo_atual.append(dados)

    def carregar_arquivo(self,nomeArquivo):
        novo_arquivo = load_workbook(f"{nomeArquivo}.xlsx")
        return novo_arquivo

    def criar_cabecalho(self, *strings):
        self.arquivo_atual.append(strings)

    def criar_nova_planilha(self, nome, posicao=0):
        self.arquivo_excel.create_sheet(nome, posicao)

    def nome_planilhas(self):
        print(self.arquivo_excel.sheetnames)

    def renomear_planilha(self, nome):
        self.arquivo_atual.title = nome

    def salvar_arquivo(self, nomeDoArquivo):
        self.arquivo_excel.save(f"{nomeDoArquivo}.xlsx")


class excel_xls():
    arquivo_excel = xlwt.Workbook()
    planilha1 = arquivo_excel.add_sheet("Planilha1")

    def add_dados(self,coluna,linha, *entrada):
        elementos = entrada

        for dado in elementos:
            self.planilha1.write(coluna, linha,dado)
            linha+=1

    def criar_cabecalho(self, coluna, linha, *entrada):
        elementos = entrada

        for item in elementos:
            self.planilha1.write(coluna,linha, item)
            linha +=1
        linhaInicial = 0

    def criar_nova_planilha(self, nome):
        self.arquivo_excel.add_sheet(f"{nome}")

    def escrever_na_celula(self,linha,coluna,valor):
        self.planilha1.write(linha,coluna,valor)

    def renomear_planilha(self, nome):
        self.planilha1.name = nome

    def salvar_arquivo(self, nome):
        self.arquivo_excel.save(f"{nome}.xls")

objexcel = excel_xlsx()

objexcel.add_dados("da","da","fgdfg")
objexcel.salvar_arquivo("werw")