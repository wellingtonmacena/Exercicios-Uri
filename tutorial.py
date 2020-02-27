#Baixar biblioteca = pip install openpyxl

#Importar openpyxl, load_workbook
from openpyxl import Workbook, load_workbook

#Criando objeto da classe Workbook
arquivo_excel = Workbook();

#Obter a planilha
planilha1 = arquivo_excel.active

#Renomear planilha
planilha1.title = "Gastos"

#Ciar uma nova planilha. Paramêtros: nome da tabela, e indice no arquivo
planilha2 = arquivo_excel.create_sheet("Ganhos", 1)

#Exibir os nome das tabelas do seu arquivo
print(arquivo_excel.sheetnames)

#Adicionando valores a planilhas nas celulas especificas
planilha1["A1"] = "Categoria"
planilha1["B1"] = "Valor"
planilha1["A2"] = "Restaurante"
planilha1["B2"] = 45.99

well = 34*43/1**7-2
planilha1["w3"] = "well"
planilha1["w2"] = well

# Paramêtros = Linha, coluna, e valor
planilha1.cell(row=14, column=1, value=43.55)


#Adicionando valores a planilhas nos indices padrão
valores = [("   ", ""),
           ("Categoria", "Valor"),
           ("Restaurante", 45.99),
           ("Transporte", 208.45),
           ("Viagem", 558.54),
           ("Video-game", 1200)]

for linha in valores:
    planilha1.append(linha)

#Adicionando formulas

planilha1["C6"] = '=SOMA(25,3)'

#Acessando o valor de uma celula
c1 = planilha1["C6"]
print(c1.value)

#Salvando uma planilha. Paramêtro: nome da planilha e extensão
arquivo_excel.save("relatorio.xlsx")

#Carregando uma planilha existente
#-- Caminho = Path onde está seu arquivo
# caminho = (r"C:\Users\wellmac\Projetos\ProjetoExcel\paraCopiar.xlsx")
# arquivo_excel = load_workbook(caminho)

#Copiando od dados de uma planilha em outra

original = arquivo_excel.get_sheet_by_name('Gastos')
copia = arquivo_excel.copy_worksheet(copia)
arquivo_excel.save('planilha.xlsx')